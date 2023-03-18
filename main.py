import Fog
import end_user
import miner
import blockchain
import random
import output
from math import ceil
import time
import modification
import new_consensus_module
import test_data
import mempool
from openpyxl import Workbook
from openpyxl import load_workbook


isblackgun = True
machineName = "[noname]"


data = modification.read_file("Sim_parameters.json")
list_of_end_users = []
fogNodes = []
transactions_list = []
list_of_authorized_miners = []
blockchainFunction = 0
blockchainPlacement = 0
number_of_miner_neighbours = data["number_of_each_miner_neighbours"]
NumOfFogNodes = data["NumOfFogNodes"]
NumOfTaskPerUser = data["NumOfTaskPerUser"]
NumOfMiners = data["NumOfMiners"]
numOfTXperBlock = data["numOfTXperBlock"]
num_of_users_per_fog_node = data["num_of_users_per_fog_node"]
blockchain_functions = ['1', '2', '3', '4']
blockchain_placement_options = ['1', '2']
expected_chain_length = ceil((num_of_users_per_fog_node * NumOfTaskPerUser * NumOfFogNodes) / numOfTXperBlock)
gossip_activated = data["Gossip_Activated"]
Automatic_PoA_miners_authorization = data["Automatic_PoA_miners_authorization?"]
Parallel_PoW_mining = data["Parallel_PoW_mining?"]
trans_delay = 0
delay_between_fog_nodes = data["delay_between_fog_nodes"]
delay_between_end_users = data["delay_between_end_users"]
poet_block_time = data['poet_block_time']
Asymmetric_key_length = data['Asymmetric_key_length']
number_of_DPoS_delegates = data['Num_of_DPoS_delegates']
user_informed = False
injectionRate = data["TX_injection_rate"] # transaction per second
queueLimit = data["QueueTooLongLimit"]
failPendingTime = data["FailPendingTime"]
uploadBandwidth = data["UploadBandwidth"] * 1024 #Bytes per second
downloadBandwidth = data["DownloadBandwidth"] * 1024 #Bytes per second

def resetSimData():
    global data, number_of_miner_neighbours, NumOfFogNodes, NumOfTaskPerUser, NumOfMiners, numOfTXperBlock, num_of_users_per_fog_node
    global expected_chain_length, gossip_activated, Automatic_PoA_miners_authorization, Parallel_PoW_mining, delay_between_fog_nodes
    global delay_between_end_users, poet_block_time, Asymmetric_key_length, number_of_DPoS_delegates, injectionRate, queueLimit
    global uploadBandwidth, downloadBandwidth, failPendingTime
    
    data = modification.read_file(machineName+"Sim_parameters.json")
    number_of_miner_neighbours = data["number_of_each_miner_neighbours"]
    NumOfFogNodes = data["NumOfFogNodes"]
    NumOfTaskPerUser = data["NumOfTaskPerUser"]
    NumOfMiners = data["NumOfMiners"]
    numOfTXperBlock = data["numOfTXperBlock"]
    num_of_users_per_fog_node = data["num_of_users_per_fog_node"]
    expected_chain_length = ceil((num_of_users_per_fog_node * NumOfTaskPerUser * NumOfFogNodes) / numOfTXperBlock)
    gossip_activated = data["Gossip_Activated"]
    Automatic_PoA_miners_authorization = data["Automatic_PoA_miners_authorization?"]
    Parallel_PoW_mining = data["Parallel_PoW_mining?"]
    delay_between_fog_nodes = data["delay_between_fog_nodes"]
    delay_between_end_users = data["delay_between_end_users"]
    poet_block_time = data['poet_block_time']
    Asymmetric_key_length = data['Asymmetric_key_length']
    number_of_DPoS_delegates = data['Num_of_DPoS_delegates'] 
    injectionRate = data["TX_injection_rate"] #per second
    queueLimit = data["QueueTooLongLimit"]
    failPendingTime = data["FailPendingTime"]
    uploadBandwidth = data["UploadBandwidth"] * 1024
    downloadBandwidth = data["DownloadBandwidth"] * 1024
    

def user_input():
    modification.initiate_files(gossip_activated)
    choose_functionality()
    choose_placement()


def choose_functionality():

    while True:
        output.choose_functionality()
        global blockchainFunction
        
        ######################################################################
        if isblackgun == True:
            if blockchainFunction == 0:
                blockchainFunction = 3 #Payment
            print("[Auto by Blackgun] " + str(blockchainFunction))
            break
        ######################################################################
        
        blockchainFunction = input()
        if blockchainFunction in blockchain_functions:
            blockchainFunction = int(blockchainFunction)
            break
        else:
            print("Input is incorrect, try again..!")


def choose_placement():
    while True:
        output.choose_placement()
        global blockchainPlacement
        
        ######################################################################
        if isblackgun == True:
            if blockchainPlacement == 0:
                blockchainPlacement = 2 #End_User
            print("[Auto by Blackgun] " + str(blockchainPlacement))
            break
        ######################################################################
        
        blockchainPlacement = input()
        if blockchainPlacement in blockchain_placement_options:
            blockchainPlacement = int(blockchainPlacement)
            break
        else:
            print("Input is incorrect, try again..!")


def initiate_network():
    for count in range(NumOfFogNodes):
        fogNodes.append(Fog.Fog(count + 1))
        for p in range(num_of_users_per_fog_node):
            list_of_end_users.append(end_user.User(p + 1, count + 1))
    output.users_and_fogs_are_up()
    if blockchainFunction == 4:
        output.GDPR_warning()
        while True:
            print("If you don't want other attributes to be added to end_users, input: done\n")
            new_attribute = input("If you want other attributes to be added to end_users, input them next:\n")
            if new_attribute == 'done':
                break
            else:
                for user in list_of_end_users:
                    user.identity_added_attributes[new_attribute] = ''
                output.user_identity_addition_reminder(len(list_of_end_users))
    for user in list_of_end_users:
        user.create_tasks(NumOfTaskPerUser, blockchainFunction, list_of_end_users)
        user.send_tasks(fogNodes)
        print("End_user " + str(user.addressParent) + "." + str(user.addressSelf) + " had sent its tasks to the fog layer")
    
    
def initiate_miners():
    the_miners_list = []

    if blockchainPlacement == 1:
        for i in range(NumOfFogNodes):
            the_miners_list.append(miner.Miner(i + 1, trans_delay, gossip_activated, uploadBandwidth, downloadBandwidth))
    if blockchainPlacement == 2:
        for i in range(NumOfMiners):
            the_miners_list.append(miner.Miner(i + 1, trans_delay, gossip_activated, uploadBandwidth, downloadBandwidth))
    for entity in the_miners_list:
        modification.write_file("temporary/" + entity.address + "_local_chain.json", {})
        miner_wallets_log_py = modification.read_file("temporary/miner_wallets_log.json")
        miner_wallets_log_py[str(entity.address)] = data['miners_initial_wallet_value']
        modification.rewrite_file("temporary/miner_wallets_log.json", miner_wallets_log_py)
    print('Miners have been initiated..')
    connect_miners(the_miners_list)
    output.miners_are_up()
    return the_miners_list


def define_trans_delay(layer):
    transmission_delay = 0
    if layer == 1:
        transmission_delay = delay_between_fog_nodes
    if layer == 2:
        transmission_delay = delay_between_end_users
    return transmission_delay


def connect_miners(miners_list):
    print("Miners will be connected in a P2P fashion now. Hold on...")
    bridges = set()
    all_components = create_components(miners_list)
    for comp in all_components:
        bridge = random.choice(tuple(comp))
        bridges.add(bridge)
    bridging(bridges, miners_list)


def bridging(bridges, miners_list):
    while len(bridges) != 1:
        bridge = random.choice(tuple(bridges))
        other_bridge = random.choice(tuple(bridges))
        same_bridge = True
        while same_bridge:
            other_bridge = random.choice(tuple(bridges))
            if other_bridge != bridge:
                same_bridge = False
        for entity in miners_list:
            if entity.address == bridge:
                entity.neighbours.add(other_bridge)
            if entity.address == other_bridge:
                entity.neighbours.add(bridge)
        bridges.remove(bridge)


def create_components(miners_list):
    all_components = set()
    for entity in miners_list:
        component = set()
        while len(entity.neighbours) < number_of_miner_neighbours:
            neighbour = random.choice(miners_list).address
            if neighbour != entity.address:
                entity.neighbours.add(neighbour)
                component.add(neighbour)
                for entity_2 in miners_list:
                    if entity_2.address == neighbour:
                        entity_2.neighbours.add(entity.address)
                        component.add(entity.address)
                        break
        if component:
            all_components.add(tuple(component))
    return all_components


def give_miners_authorization(the_miners_list, the_type_of_consensus):
    if the_type_of_consensus == 1:
        wanted, float_portion = output.AI_assisted_mining_wanted()
        if wanted:
            num_of_miners_requested_to_use_AI = ceil(float_portion * len(the_miners_list))
            num_of_miners_instructed_to_use_AI = 0
            while num_of_miners_instructed_to_use_AI < num_of_miners_requested_to_use_AI:
                random_miner = random.choice(the_miners_list)
                if not random_miner.adversary:
                    random_miner.adversary = True
                    num_of_miners_instructed_to_use_AI += 1
            print(str(num_of_miners_instructed_to_use_AI) + ' miners were successfully instructed to use AI.')
        return wanted
    if the_type_of_consensus == 3:
        # automated approach:
        if Automatic_PoA_miners_authorization:
            for i in range(len(the_miners_list)):
                the_miners_list[i].isAuthorized = True
                list_of_authorized_miners.append(the_miners_list[i])
        else:
            # user input approach:
            output.authorization_trigger(blockchainPlacement, NumOfFogNodes, NumOfMiners)
            while True:
                authorized_miner = input()
                if authorized_miner == "done":
                    break
                else:
                    for node in the_miners_list:
                        if node.address == "Miner_" + authorized_miner:
                            node.isAuthorized = True
                            list_of_authorized_miners.append(node)
    return None


def initiate_genesis_block(AI_wanted):
    genesis_transactions = ["genesis_block"]
    for i in range(len(miner_list)):
        genesis_transactions.append(miner_list[i].address)
    genesis_block = new_consensus_module.generate_new_block(genesis_transactions, 'The Network', 0, type_of_consensus, AI_wanted, False)
    output.block_info(genesis_block, type_of_consensus)
    for elem in miner_list:
        elem.receive_new_block(genesis_block, type_of_consensus, miner_list, blockchainFunction)
    output.genesis_block_generation()


def send_tasks_to_BC():
    global user_informed
    for node in fogNodes:
        node.send_tasks_to_BC(user_informed, isblackgun)
        if not user_informed:
            user_informed = True


def store_fog_data():
    for node in fogNodes:
        log = open('temporary/Fog_node_'+str(node.address)+'.txt', 'w')
        log.write(str(node.local_storage))



def inform_miners_of_users_wallets():
    if blockchainFunction == 3:
        user_wallets = {}
        for user in list_of_end_users:
            wallet_info = {'parent': user.addressParent,
                           'self': user.addressSelf,
                           'wallet_value': user.wallet}
            user_wallets[str(user.addressParent) + '.' + str(user.addressSelf)] = wallet_info
        for i in range(len(miner_list)):
            modification.rewrite_file(str("temporary/" + miner_list[i].address + "_users_wallets.json"), user_wallets)

import sys
if __name__ == '__main__':
    # 
    if len(sys.argv) == 2:
        isblackgun = bool(sys.argv[1])
        blockchainFunction = 3 #Payment
        blockchainPlacement = 2 #End_User
        num_of_consensus = 2 #PoS
        print("[blackgun auto mode1]" + str(isblackgun)) 
    
    if len(sys.argv) == 5:
        isblackgun = bool(sys.argv[1])
        blockchainFunction = int(sys.argv[2])
        blockchainPlacement = int(sys.argv[3])
        num_of_consensus = int(sys.argv[4])
        print("[blackgun auto mode2]" + str(isblackgun)) 
        
    if len(sys.argv) >= 6:
        isblackgun = bool(sys.argv[1])
        machineName = sys.argv[2]
        blockchainFunction = int(sys.argv[3])
        blockchainPlacement = int(sys.argv[4])
        num_of_consensus = int(sys.argv[5])
        blockchain.setPrefix(sys.argv[2])
        resetSimData()
        print("[blackgun auto mode3]" + str(isblackgun)) 
        if len(sys.argv) > 6:
            numOfTXperBlock = int(sys.argv[6])
            injectionRate = int (sys.argv[7])
            print("[blackgun auto mode3+] overwrite: " + "tx per block = " + str(numOfTXperBlock) + ", injection rate = " + str(injectionRate))
    

    
    user_input()
    initiate_network()
    type_of_consensus = new_consensus_module.choose_consensus(isblackgun, num_of_consensus)
    trans_delay = define_trans_delay(blockchainPlacement)
    miner_list = initiate_miners()
    AI_assisted_mining_wanted = give_miners_authorization(miner_list, type_of_consensus)
    inform_miners_of_users_wallets()
    blockchain.stake(miner_list, type_of_consensus)
    initiate_genesis_block(AI_assisted_mining_wanted)
    send_tasks_to_BC()
    
    #print("++++++++++++++++++++++++++++ time start")
    time_start = time.time()
    if blockchainFunction == 2:
        expected_chain_length = ceil((num_of_users_per_fog_node * NumOfTaskPerUser * NumOfFogNodes) / numOfTXperBlock)
    #print("++++++++++++++++++++++++++++ A:"+str(time.time() - time_start))
    
    new_consensus_module.miners_trigger(miner_list, type_of_consensus, expected_chain_length, Parallel_PoW_mining,
                                        numOfTXperBlock, blockchainFunction, poet_block_time, Asymmetric_key_length,
                                        number_of_DPoS_delegates, AI_assisted_mining_wanted, injectionRate, queueLimit, failPendingTime)
    print("totalBlockTime = " + str(test_data.totalBlockTime))
    
    #print("++++++++++++++++++++++++++++ B:"+str(time.time() - time_start))
    blockchain.award_winning_miners(len(miner_list), miner_list)
    
    blockchain.fork_analysis(miner_list)
    output.finish()
    store_fog_data()
    
    averageDownloadDataUsage = 0
    averageUploadDataUsage = 0
    for m in miner_list:
        print(str(m.address) + " upload: " + str(m.uploadDataUsage) + " download: " + str(m.downloadDataUsage) + "(Byte)")
        averageDownloadDataUsage += m.downloadDataUsage
        averageUploadDataUsage += m.uploadDataUsage
    averageUploadDataUsage /= NumOfMiners
    averageDownloadDataUsage /= NumOfMiners
    
    elapsed_time = time.time() - time_start
    #print("++++++++++++++++++++++++++++ time end")
    
    number_of_user = NumOfFogNodes * num_of_users_per_fog_node
    number_of_TX = number_of_user * NumOfTaskPerUser
    #number_of_block = ceil(number_of_TX / numOfTXperBlock)
    number_of_block = test_data.numOfBlock
    print("number of block = " + str(number_of_block))
    
    print("totalBlockTime = " + str(test_data.totalBlockTime))
    average_block_time = test_data.totalBlockTime / (float)(number_of_block)
    average_block_time_ms = average_block_time * 1000.0
    average_transaction_pending_time_ms = test_data.totalTXPendingTime / (float)(test_data.totalTXNum) * 1000.0
    
    print("elapsed time = " + str(elapsed_time) + " seconds")
    print("[BG] average block time = " + str(average_block_time_ms) + " ms")
    with open(machineName + 'result_log.txt', 'a+') as resultfile:
        resultfile.write("No. user: " + str(number_of_user))
        resultfile.write(" , No. miner: " + str(NumOfMiners))
        resultfile.write(" , No. minerNeighbours: " + str(number_of_miner_neighbours))
        resultfile.write(" , No. tx: " + str(number_of_TX))
        resultfile.write(" , No. block: " + str(number_of_block))
        resultfile.write(" , average block time(secs): " + str(average_block_time))
        resultfile.write(" , average upload data (bytes): " + str(averageUploadDataUsage))
        resultfile.write(" , average download data (bytes): " + str(averageDownloadDataUsage))
        resultfile.write(" , elapsed time(secs): " + str(elapsed_time) + "\n")
    with open(machineName + 'result_avrTime.txt', 'a+') as resultTimeFile:
        resultTimeFile.write(str(average_block_time)+"\n")
    
    print("TX_injection_rate = " + str(injectionRate))
    
    filename = machineName + 'result.xlsx'
    new_row = [type_of_consensus, blockchainFunction, blockchainPlacement, number_of_user, NumOfMiners, number_of_miner_neighbours, number_of_TX, delay_between_fog_nodes, delay_between_end_users, uploadBandwidth / 1024, downloadBandwidth / 1024, gossip_activated, injectionRate, numOfTXperBlock, '<-parameter / result->', number_of_block, test_data.failTime, average_transaction_pending_time_ms, average_block_time_ms, test_data.totalBlockTime, test_data.totalBlockPrepareTime, test_data.totalUploadTime, test_data.totalDownloadTime, test_data.totalNetworkDelayTime, elapsed_time, averageUploadDataUsage, averageDownloadDataUsage]
    headers_row = ['consensus', 'function', 'placement', 'No. user', 'No. miner', 'No. minerNeighbours', 'init No. tx:', 'delay between fog node(ms)', 'delay between end users(ms)', 'upload bandwidth(KB/S)', 'download bandwidth(KB/S)', 'gossip', 'injection rate(per sec)', 'tx per block', '<-parameter / result->', 'final No. block', 'fail time(secs)', 'average transaction pending time(ms)', 'average block time(ms)', 'simulation time(sec)', 'total prepare time(sec)', 'total upload time(sec)', 'total download time(sec)', 'total network delay time(sec)', 'elapsed time(secs)', 'average upload data(bytes)', 'average download data(bytes)']
    
    # Confirm file exists. 
    # If not, create it, add headers, then append new data
    try:
        wb = load_workbook(filename)
        ws = wb.worksheets[0]  # select first worksheet
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(headers_row)

    ws.append(new_row)


    
    for col_num, value in enumerate(headers_row, start=1):
        ws.cell(row=1, column=col_num, value=value)
    
    new_row = ws[ws.max_row]

    new_row[-11].number_format = '0.000000'
    new_row[-10].number_format = '0.000000'
    new_row[-9].number_format = '0.000000'
    new_row[-8].number_format = '0.000000'
    new_row[-7].number_format = '0.000000'
    new_row[-6].number_format = '0.000000'
    new_row[-5].number_format = '0.000000'
    new_row[-4].number_format = '0.000000'
    new_row[-3].number_format = '0.000'
    new_row[-2].number_format = '0.00'
    new_row[-1].number_format = '0.00'
    
    wb.save(filename)
    
    mempool.MemPool.close()
    